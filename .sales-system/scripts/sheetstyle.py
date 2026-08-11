#!/usr/bin/env python3
"""
sheetstyle.py — the Excel presentation layer for sales registries.

The goal is that opening a registry in Excel feels like opening a finished tool, not a
data dump: header frozen and filtered, columns the right width, dates and money
formatted, picklist columns offering dropdowns of the org's real values, and colour
where it carries meaning.

Two things worth knowing about the design.

First, the styling contract is *declarative and idempotent*. It's derived from the
schema, not hand-tuned per file, and reapplying it is always safe. That's what keeps
the file readable after weeks of edits.

Second, writes update in place rather than regenerating. Any manual formatting a person
adds to their own rows survives. The contract is re-asserted around it.

Used by csvguard.py; not normally called directly, though `--restyle` is handy after
someone has mangled a sheet.
"""

import json
import os
import re
import sys
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.comments import Comment
    from openpyxl.formatting.formatting import ConditionalFormattingList
except ImportError:  # pragma: no cover
    print("error: openpyxl is required for .xlsx registries.\n"
          "  pip install openpyxl --break-system-packages", file=sys.stderr)
    raise

DATA_SHEET = "Data"
LOOKUP_SHEET = "_Lists"        # holds picklist values for dropdowns; hidden

# --------------------------------------------------------------- colour util

def _hex(c):
    c = (c or "").lstrip("#").strip()
    if len(c) == 8:          # openpyxl sometimes hands back AARRGGBB
        c = c[2:]
    return c.upper() if len(c) == 6 else "1F3A5F"


def _rgb(c):
    c = _hex(c)
    return tuple(int(c[i:i + 2], 16) for i in (0, 2, 4))


def _hx(rgb):
    return "".join(f"{max(0, min(255, int(round(v)))):02X}" for v in rgb)


def mix(c, other, amount):
    """Blend `c` toward `other` by amount 0..1."""
    a, b = _rgb(c), _rgb(other)
    return _hx([a[i] + (b[i] - a[i]) * amount for i in range(3)])


def tint(c, amount):
    """Lighten toward white — how every fill in the palette is derived."""
    return mix(c, "FFFFFF", amount)


def shade(c, amount):
    return mix(c, "000000", amount)


def luminance(c):
    r, g, b = [v / 255 for v in _rgb(c)]
    f = lambda v: v / 12.92 if v <= 0.03928 else ((v + 0.055) / 1.055) ** 2.4
    return 0.2126 * f(r) + 0.7152 * f(g) + 0.0722 * f(b)


def readable_on(bg):
    """Pick black or white text for a background. Without this, a dark brand colour
    with dark text produces an unreadable header and the whole thing looks broken."""
    return "FFFFFF" if luminance(bg) < 0.45 else "1A1A1A"


def readable_ink(c, bg="FFFFFF"):
    """Darken a colour until it's legible as text on `bg`. Brand colours are chosen to
    look good as large blocks, not as 10pt type."""
    out = _hex(c)
    for _ in range(14):
        l1, l2 = luminance(out) + 0.05, luminance(bg) + 0.05
        if (max(l1, l2) / min(l1, l2)) >= 4.5:
            return out
        out = shade(out, 0.12)
    return out


# ------------------------------------------------------------------ palette
# Derived from the org's brand colours so the workbooks look like the company's,
# then muted hard. A sheet where everything shouts is a sheet nobody reads.

DEFAULT_BRAND = {"primary": "1F3A5F", "secondary": "3E6E8E",
                 "accent": "C86A3B", "neutral": "5B6472", "tint_strength": 0.90}


def load_brand(root):
    if root:
        p = os.path.join(root, ".sales-system", "brand.json")
        if os.path.exists(p):
            try:
                with open(p, encoding="utf-8") as f:
                    b = dict(DEFAULT_BRAND)
                    b.update({k: v for k, v in json.load(f).items()
                              if k in DEFAULT_BRAND and v})
                    return b
            except (json.JSONDecodeError, OSError):
                pass
    return dict(DEFAULT_BRAND)


class Palette:
    """Everything the sheet needs, derived from four brand colours."""

    def __init__(self, brand):
        p = _hex(brand["primary"])
        s = _hex(brand["secondary"])
        a = _hex(brand["accent"])
        n = _hex(brand["neutral"])
        t = float(brand.get("tint_strength", 0.90))

        # A near-black primary makes a heavy, funereal header. Lift it toward the
        # secondary so the bar reads as brand rather than as a redaction.
        self.header_bg = mix(p, s, 0.35) if luminance(p) < 0.06 else p
        self.header_fg = readable_on(self.header_bg)
        self.header_rule = shade(self.header_bg, 0.25)

        self.band = tint(s, min(t + 0.06, 0.975))    # zebra stripe, barely there
        self.key_bg = tint(s, t)                     # id column
        self.derived_bg = tint(n, 0.955)             # computed columns
        self.derived_ink = readable_ink(n)
        self.rule = tint(n, 0.80)                    # hairline
        self.accent = a
        self.accent_bg = tint(a, t)
        self.accent_ink = readable_ink(a)
        self.tab = self.header_bg

        # Risk colours are deliberately NOT brand-derived. Red must read as red
        # whatever the company's palette is, or the signal stops working.
        self.red_bg, self.red_ink = "FBE3E1", "A32C22"
        self.amber_bg, self.amber_ink = "FDF0D9", "8A5A00"
        self.green_bg, self.green_ink = "E4F1E6", "1E6B34"
        self.info_bg, self.info_ink = tint(s, 0.88), readable_ink(s)
        self.mute_bg, self.mute_ink = tint(n, 0.92), tint(n, 0.25)


def cf_fill(colour):
    """Fill for a conditional-formatting rule. Excel reads bgColor here; fgColor is
    silently ignored, which produces a rule that exists but does nothing visible."""
    return PatternFill(start_color=colour, end_color=colour, fill_type="solid",
                       bgColor=colour)


def build_styles(pal):
    return {
        "header_fill": PatternFill("solid", fgColor=pal.header_bg),
        "header_font": Font(color=pal.header_fg, bold=True, size=10,
                            name="Calibri"),
        "band_fill": cf_fill(pal.band),
        "key_fill": PatternFill("solid", fgColor=pal.key_bg),
        "derived_fill": PatternFill("solid", fgColor=pal.derived_bg),
        "derived_font": Font(color=pal.derived_ink, size=10, italic=True),
        "body_font": Font(color="1A1A1A", size=10),
        # Horizontal hairlines only. Vertical lines on a 45-column sheet make a
        # cage; whitespace separates columns perfectly well.
        "row_border": Border(bottom=Side(style="thin", color=pal.rule)),
        "header_border": Border(bottom=Side(style="medium", color=pal.header_rule)),
        # NB: conditional-formatting fills must set bgColor. A dxf built with fgColor
        # is written to the file but renders as no fill in Excel.
        "red": (cf_fill(pal.red_bg), Font(color=pal.red_ink, bold=True)),
        "amber": (cf_fill(pal.amber_bg), Font(color=pal.amber_ink)),
        "green": (cf_fill(pal.green_bg), Font(color=pal.green_ink)),
        "info": (cf_fill(pal.info_bg), Font(color=pal.info_ink, bold=True)),
        "mute": (cf_fill(pal.mute_bg), Font(color=pal.mute_ink, italic=True)),
    }

# Semantic colouring by picklist value -> style key. Matching is case-insensitive and
# covers the common vocabularies; an org's bespoke values fall through uncoloured rather
# than being guessed at, which is the safe failure.
SEMANTIC = {
    "red": "red", "at risk": "red", "churned": "red", "closed lost": "red",
    "disqualified": "red", "loss": "red", "do not contact": "red", "junk": "red",
    "blocked": "red", "conflict": "red", "high": "red", "contraction": "red",
    "failed": "red", "rumour": "red", "overdue": "red",

    "yellow": "amber", "delayed": "amber", "awaiting approval": "amber",
    "pending-push": "amber", "on hold": "amber", "nurture": "amber", "snoozed": "amber",
    "medium": "amber", "not started": "amber", "reported": "amber", "downgraded": "amber",
    "extended": "amber", "warm": "amber",

    "green": "green", "closed won": "green", "renewed": "green", "qualified": "green",
    "done": "green", "synced": "green", "expansion": "green", "confirmed": "green",
    "customer or partner": "green", "commit": "green", "auto-renewed": "green",
    "acted on": "green",

    "replied to sequence": "info", "conversation started": "info", "opp created": "info",
    "drafted": "info", "in progress": "info", "scheduled": "info", "hot": "info",
    "in negotiation": "info", "negotiate": "info", "proposal": "info",

    "cancelled": "mute", "dismissed": "mute", "not required": "mute", "omitted": "mute",
    "local-only": "mute", "low": "mute", "cold": "mute", "watching": "mute",
    "passed to partner": "mute",
}

NUMBER_FORMATS = {
    "date": "yyyy-mm-dd;;\"—\"",
    "money": '#,##0;[Red](#,##0);"—"',
    "number": '#,##0.##;-#,##0.##;"—"',
    "bool": "@",
    "id": "@",
    "text": "@",
    "enum": "@",
    "email": "@",
    "url": "@",
}

# Columns whose content is long prose — wider, wrapped, and pushed right in the sheet
# so they don't force horizontal scrolling past the fields people actually scan.
WIDE_HINTS = ("notes", "why", "so_what", "description", "pain", "next_step",
              "decision_process", "paper_process", "action_taken", "evidence",
              "completion_evidence", "risk_flags", "close_plan_gaps",
              "churn_risk_reason", "expansion_signal", "loss_notes", "headline",
              "contact_restrictions", "trigger", "sources")


def is_prose(col):
    """Wide, wrapped columns. Type wins over name: `next_step` is prose, but
    `next_step_date` is a date and must stay narrow."""
    if col.get("type", "text") not in ("text", "enum"):
        return False
    return any(h in col["name"] for h in WIDE_HINTS)


def col_width(col):
    n, t = col["name"], col.get("type", "text")
    if is_prose(col):
        return 42
    if t in ("date",):
        return 12
    if t in ("money", "number"):
        return 12
    if t == "bool":
        return 8
    if t == "id":
        return 12
    if t == "enum":
        longest = max((len(v) for v in col.get("values", [])), default=12)
        return min(max(longest + 4, 12), 26)
    if t in ("email", "url"):
        return 26
    return min(max(len(n.replace("_", " ")) + 6, 14), 30)


def is_derived(col):
    note = (col.get("note") or "").upper()
    return col.get("derived") is True or note.startswith("DERIVED") or "DERIVED" in note[:40]


# ------------------------------------------------------------------- coercion

def to_excel_value(raw, ctype):
    """Canonical string -> a native Excel value, so Excel sorts and filters properly.

    This is the actual fix for the CSV round-trip problem: a real date cell can't be
    reformatted into ambiguity, and a real number can't acquire a currency symbol."""
    s = (raw or "").strip()
    if s == "":
        return None
    if ctype == "date":
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return s
    if ctype in ("money", "number"):
        try:
            f = float(s)
            return int(f) if f == int(f) else f
        except ValueError:
            return s
    return s


def from_excel_value(v, ctype):
    """Native Excel value -> the canonical string the schema expects."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    if isinstance(v, bool):
        return "yes" if v else "no"
    return str(v).strip()


# ------------------------------------------------------------------------ io

def read_xlsx(path, schema=None):
    """Return (header, rows) as canonical strings, matching the CSV reader's contract."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    raw = [str(h).strip() for h in rows[0] if h is not None]
    # headers are displayed Title Case; map them back to schema field names
    known = {c["name"]: c for c in (schema or {}).get("columns", [])}
    pretty = {n.replace("_", " ").title(): n for n in known}
    header = [pretty.get(h, h.replace(" ", "_").lower() if h in pretty else h) for h in raw]
    header = [pretty.get(h, h) for h in raw]
    spec = {c["name"]: c.get("type", "text") for c in (schema or {}).get("columns", [])}
    body = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        out = []
        for i, h in enumerate(header):
            v = r[i] if i < len(r) else None
            out.append(from_excel_value(v, spec.get(h, "text")))
        body.append(out)
    return header, body


def write_xlsx(path, header, rows, schema, picklists=None, brand=None):
    """Write values, then assert the styling contract.

    Updates in place when the file exists so a person's own highlighting survives."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    exists = os.path.exists(path)
    if exists:
        wb = openpyxl.load_workbook(path)
        ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.worksheets[0]
        ws.title = DATA_SHEET
        # clear validations; they're re-added from the current schema below
        ws.data_validations.dataValidation = []
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = DATA_SHEET

    spec = {c["name"]: c for c in schema["columns"]}
    old_rows = ws.max_row

    # NB: openpyxl's ws.cell(row, col, value=X) only assigns when X is not None — passing
    # None is a silent no-op that leaves whatever was there. Writing a blank over an
    # existing cell therefore has to go through .value, or old data survives in place.
    # This bit us for real: inserting columns mid-schema left the previous occupants
    # sitting under the new headers, duplicating data with no error.
    old_cols = ws.max_column
    for ci, name in enumerate(header, start=1):
        ws.cell(row=1, column=ci).value = name
    for ri, row in enumerate(rows, start=2):
        for ci, name in enumerate(header, start=1):
            c = spec.get(name, {"type": "text"})
            ws.cell(row=ri, column=ci).value = to_excel_value(
                row[ci - 1] if ci - 1 < len(row) else "", c.get("type", "text"))

    # shrink: clear rows and columns that no longer carry data
    for ri in range(len(rows) + 2, old_rows + 1):
        for ci in range(1, max(old_cols, len(header)) + 1):
            ws.cell(row=ri, column=ci).value = None
    for ci in range(len(header) + 1, old_cols + 1):
        for ri in range(1, max(old_rows, len(rows) + 1) + 1):
            ws.cell(row=ri, column=ci).value = None

    if brand is None:
        brand = load_brand(find_root(path))
    apply_contract(wb, ws, header, len(rows), schema, picklists, brand)

    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)


# ------------------------------------------------------- the styling contract

# The columns worth keeping in view when someone scrolls right through 45 fields.
# Freezing these is the single biggest usability win on a wide registry — without it
# you're looking at a number with no idea whose it is.
ANCHOR_CANDIDATES = ("id", "name", "title", "account_name", "company", "last_name", "term")


def anchor_span(header):
    n = 0
    for h in header[:3]:
        if h in ANCHOR_CANDIDATES:
            n += 1
        else:
            break
    return min(n, 2)


def apply_contract(wb, ws, header, n_rows, schema, picklists=None, brand=None):
    """Idempotent. Safe to reapply any number of times."""
    pal = Palette(brand or DEFAULT_BRAND)
    st = build_styles(pal)
    spec = {c["name"]: c for c in schema["columns"]}
    last_row = max(n_rows + 1, 2)
    last_col = len(header)
    last_letter = get_column_letter(last_col)

    # Excel's own gridlines fight every other visual decision on the sheet. Off.
    # Rules accumulate across writes, and the stale ones hold higher priority — an
    # early rule written when the sheet was empty covers one row and silently wins.
    ws.conditional_formatting = ConditionalFormattingList()

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_properties.tabColor = pal.tab

    # --- header
    for ci, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value = name.replace("_", " ").title() if name in spec else name
        cell.fill = st["header_fill"]
        cell.font = st["header_font"]
        cell.border = st["header_border"]
        cell.alignment = Alignment(vertical="center", horizontal="left",
                                   wrap_text=False, indent=1)
        col = spec.get(name)
        note = (col or {}).get("note")
        if note:
            # schema guidance lives where the person needs it: on the header
            cell.comment = Comment(note[:800], "sales system", height=110, width=320)
        ws.column_dimensions[get_column_letter(ci)].width = col_width(
            col or {"name": name, "type": "text"})
    ws.row_dimensions[1].height = 30

    anchors = anchor_span(header)
    ws.freeze_panes = f"{get_column_letter(anchors + 1)}2" if anchors else "A2"
    ws.auto_filter.ref = f"A1:{last_letter}{last_row}"

    # --- body
    for ci, name in enumerate(header, start=1):
        col = spec.get(name, {"name": name, "type": "text"})
        ctype = col.get("type", "text")
        fmt = NUMBER_FORMATS.get(ctype, "@")
        wide = is_prose(col)
        derived = is_derived(col)
        is_key = (ci <= anchors)
        for ri in range(2, last_row + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.number_format = fmt
            cell.border = st["row_border"]
            cell.font = st["derived_font"] if derived else st["body_font"]
            cell.alignment = Alignment(
                vertical="center",
                horizontal="right" if ctype in ("money", "number") else "left",
                wrap_text=wide, indent=0 if ctype in ("money", "number") else 1)
            if is_key:
                cell.fill = st["key_fill"]
            elif derived:
                cell.fill = st["derived_fill"]
        ws.row_dimensions[1].height = 30

    for ri in range(2, last_row + 1):
        ws.row_dimensions[ri].height = 20

    # --- per-column dropdowns and semantic colour.
    # Added BEFORE the zebra band so they hold higher priority; Excel applies rules in
    # priority order and the first to set a fill wins, so banding added first would
    # swallow every status colour on even rows.
    for ci, name in enumerate(header, start=1):
        col = spec.get(name, {"name": name, "type": "text"})
        letter = get_column_letter(ci)
        values = resolve_values(col, schema, picklists)
        if col.get("type") == "enum" and values:
            add_dropdown(wb, ws, letter, last_row, name, values)
            add_semantic_rules(ws, letter, last_row, values, st)
        if col.get("type") == "bool":
            add_bool_rules(ws, letter, last_row, st)

    add_date_rules(ws, header, last_row, spec, st)

    # Zebra banding last, so it fills only the rows no semantic rule claimed. Done as a
    # rule rather than baked fills so it survives sorting and row insertion.
    ws.conditional_formatting.add(f"A2:{last_letter}{last_row}", FormulaRule(
        formula=["MOD(ROW(),2)=0"], fill=st["band_fill"], stopIfTrue=False))

    # Land the cursor on the first real cell rather than wherever it was left.
    ws.sheet_view.selection[0].activeCell = "A2"
    ws.sheet_view.selection[0].sqref = "A2"


def resolve_values(col, schema, picklists):
    """Profile overrides win — dropdowns must offer what the CRM actually accepts."""
    if picklists:
        reg = picklists.get(schema["registry"], {})
        if col["name"] in reg and col["name"] in schema.get("picklists_overridable", []):
            return list(reg[col["name"]])
    return list(col.get("values", []))


def add_dropdown(wb, ws, letter, last_row, name, values):
    """Excel caps an inline list at 255 chars, so values go on a hidden sheet and the
    validation points at the range. Also means the list is visible and editable if
    someone needs to check what's allowed."""
    if LOOKUP_SHEET in wb.sheetnames:
        lut = wb[LOOKUP_SHEET]
    else:
        lut = wb.create_sheet(LOOKUP_SHEET)
        lut.sheet_state = "hidden"

    target_col = None
    for ci in range(1, lut.max_column + 2):
        if lut.cell(row=1, column=ci).value in (name, None):
            target_col = ci
            break
    L = get_column_letter(target_col)
    for ri in range(1, lut.max_row + 2):
        lut.cell(row=ri, column=target_col, value=None)
    lut.cell(row=1, column=target_col, value=name)
    for i, v in enumerate(values, start=2):
        lut.cell(row=i, column=target_col, value=v)

    ref = f"'{LOOKUP_SHEET}'!${L}$2:${L}${len(values) + 1}"
    dv = DataValidation(type="list", formula1=ref, allow_blank=True, showDropDown=False)
    dv.error = ("That isn't one of the allowed values for this column. Pick from the "
                "dropdown, or if your CRM has a value the system doesn't know about, "
                "ask to refresh the CRM profile.")
    dv.errorTitle = "Not an allowed value"
    dv.prompt = "Pick from the list"
    dv.promptTitle = name.replace("_", " ").title()
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{last_row}")


def add_semantic_rules(ws, letter, last_row, values, st):
    rng = f"{letter}2:{letter}{last_row}"
    for v in values:
        key = SEMANTIC.get(v.strip().lower())
        if not key:
            continue
        fill, font = st[key]
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'EXACT({letter}2,"{v}")'], fill=fill, font=font, stopIfTrue=True))


def add_bool_rules(ws, letter, last_row, st):
    rng = f"{letter}2:{letter}{last_row}"
    fill, font = st["mute"]
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'EXACT({letter}2,"no")'], fill=fill, font=font, stopIfTrue=True))


def add_date_rules(ws, header, last_row, spec, st):
    """Overdue dates go red — but only where being in the past actually means something.

    A created_date in the past is normal; a due_date in the past is a problem. Colouring
    both teaches people to ignore the colour."""
    meaningful = ("due_date", "next_action_date", "next_step_date", "close_date",
                  "renewal_date", "contract_end_date", "notice_deadline",
                  "conversation_target_date", "snooze_until", "poc_end_date")
    red_fill, red_font = st["red"]
    amber_fill, amber_font = st["amber"]
    for ci, name in enumerate(header, start=1):
        if spec.get(name, {}).get("type") != "date":
            continue
        if not any(name == m or name.endswith(m) for m in meaningful):
            continue
        letter = get_column_letter(ci)
        rng = f"{letter}2:{letter}{last_row}"
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({letter}2<>"",{letter}2<TODAY())'],
            fill=red_fill, font=red_font, stopIfTrue=True))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND({letter}2<>"",{letter}2<=TODAY()+7)'],
            fill=amber_fill, font=amber_font, stopIfTrue=True))


def find_root(start):
    p = os.path.abspath(start)
    if os.path.isfile(p):
        p = os.path.dirname(p)
    while True:
        if os.path.isdir(os.path.join(p, ".sales-system")):
            return p
        parent = os.path.dirname(p)
        if parent == p:
            return None
        p = parent


def restyle(path, schema, picklists=None, brand=None):
    header, rows = read_xlsx(path, schema)
    write_xlsx(path, header, rows, schema, picklists, brand)
    return len(rows)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        print("usage: sheetstyle.py --restyle <file.xlsx> [--project <root>]")
        sys.exit(2)
    if sys.argv[1] == "--restyle":
        p = sys.argv[2]
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        import csvguard
        s, root = csvguard.schema_for_file(p)
        if not s:
            print(f"error: no schema matches {p}", file=sys.stderr)
            sys.exit(2)
        n = restyle(p, s, csvguard.load_picklist_overrides(root))
        print(f"restyled {os.path.basename(p)} ({n} rows)")


# ------------------------------------------------------------------ preview
# Choosing brand colours by editing brand.json, restyling, and opening Excel is a slow
# loop. This renders the same palette to HTML so the choice can be made in seconds.

def preview_html(path, schema, picklists=None, brand=None, out=None, limit=12):
    brand = brand or load_brand(find_root(path))
    pal = Palette(brand)
    header, rows = read_xlsx(path, schema) if path.lower().endswith(".xlsx") else ([], [])
    spec = {c["name"]: c for c in schema["columns"]}

    show = [h for h in header if h in spec and (
        h in ("id", "name", "account_name", "stage", "amount", "close_date",
              "next_step", "next_step_date", "health", "owner", "risk_flags",
              "sync_status", "status", "due_date", "title", "company", "priority",
              "contract_end_date", "current_value", "so_what", "headline")
    )][:11]
    if not show:
        show = header[:9]
    idx = [header.index(h) for h in show]
    anchors = anchor_span(header)

    def sem(name, val):
        if not val:
            return None
        col = spec.get(name, {})
        if col.get("type") == "bool" and val.lower() == "no":
            return "mute"
        if col.get("type") != "enum":
            return None
        return SEMANTIC.get(val.strip().lower())

    css_sem = {
        "red": (pal.red_bg, pal.red_ink, 700), "amber": (pal.amber_bg, pal.amber_ink, 400),
        "green": (pal.green_bg, pal.green_ink, 400), "info": (pal.info_bg, pal.info_ink, 700),
        "mute": (pal.mute_bg, pal.mute_ink, 400),
    }
    today = date.today().isoformat()
    overdue_cols = ("due_date", "next_step_date", "close_date", "next_action_date",
                    "contract_end_date", "notice_deadline", "conversation_target_date")

    th = "".join(
        f'<th style="padding:9px 12px;text-align:'
        f'{"right" if spec.get(h,{}).get("type") in ("money","number") else "left"};'
        f'white-space:nowrap">{h.replace("_"," ").title()}</th>' for h in show)

    body = []
    for ri, r in enumerate(rows[:limit]):
        band = f"background:#{pal.band}" if ri % 2 else "background:#FFFFFF"
        tds = []
        for n, i in zip(show, idx):
            v = r[i] if i < len(r) else ""
            col = spec.get(n, {})
            ctype = col.get("type", "text")
            style = (f"padding:8px 12px;border-bottom:1px solid #{pal.rule};"
                     f"font-size:12.5px;color:#1A1A1A;")
            if ctype in ("money", "number"):
                style += "text-align:right;font-variant-numeric:tabular-nums;"
                if v:
                    try:
                        v = f"{float(v):,.0f}"
                    except ValueError:
                        pass
            if show.index(n) < anchors:
                style += f"background:#{pal.key_bg};font-weight:600;"
            if is_derived(col):
                style += f"background:#{pal.derived_bg};color:#{pal.derived_ink};font-style:italic;"
            inner = v or '<span style="opacity:.35">—</span>'
            k = sem(n, v)
            if k:
                bg, ink, w = css_sem[k]
                style += f"background:#{bg};color:#{ink};font-weight:{w};"
            elif any(n.endswith(m) for m in overdue_cols) and v and v < today:
                style += f"background:#{pal.red_bg};color:#{pal.red_ink};font-weight:700;"
            tds.append(f'<td style="{style}">{inner}</td>')
        body.append(f'<tr style="{band}">' + "".join(tds) + "</tr>")

    swatches = "".join(
        f'<div style="display:flex;align-items:center;gap:8px">'
        f'<span style="width:26px;height:26px;border-radius:6px;background:#{c};'
        f'border:1px solid rgba(0,0,0,.12);display:inline-block"></span>'
        f'<code style="font-size:11px;color:#555">{lbl} #{c}</code></div>'
        for lbl, c in [("primary", _hex(brand["primary"])),
                       ("secondary", _hex(brand["secondary"])),
                       ("accent", _hex(brand["accent"])),
                       ("header", pal.header_bg), ("band", pal.band), ("key", pal.key_bg)])

    html = f"""<!doctype html><meta charset="utf-8">
<title>{schema['registry']} — styling preview</title>
<body style="margin:0;padding:28px;background:#F4F5F7;
 font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif">
<div style="max-width:1180px;margin:0 auto">
 <h1 style="font-size:17px;margin:0 0 4px;color:#1A1A1A">
   {schema['registry'].replace('_',' ').title()} — how this looks in Excel</h1>
 <p style="margin:0 0 18px;color:#666;font-size:13px">
   Palette derived from brand.json. Edit those colours, rerun preview, and this updates.</p>
 <div style="display:flex;gap:18px;flex-wrap:wrap;margin-bottom:18px;padding:12px 14px;
   background:#fff;border-radius:10px;border:1px solid #E3E6EA">{swatches}</div>
 <div style="background:#fff;border-radius:10px;overflow:hidden;
   border:1px solid #E3E6EA;box-shadow:0 1px 3px rgba(0,0,0,.06)">
  <div style="height:4px;background:#{pal.accent}"></div>
  <div style="overflow-x:auto"><table style="border-collapse:collapse;width:100%">
   <thead><tr style="background:#{pal.header_bg};color:#{pal.header_fg};
     font-size:11.5px;font-weight:700;letter-spacing:.03em">{th}</tr></thead>
   <tbody>{''.join(body)}</tbody></table></div>
 </div>
 <p style="color:#777;font-size:11.5px;margin-top:14px">
  First {anchors} column(s) frozen · gridlines off · zebra banding · overdue dates flagged ·
  picklist columns carry dropdowns of your CRM's real values (not shown here).
  Colour fills the whole cell — Excel has no way to round a background inside a cell,
  so this preview shows exactly what the workbook does.</p>
</div></body>"""
    out = out or os.path.join(os.path.dirname(path), f"_preview-{schema['registry']}.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    return out
